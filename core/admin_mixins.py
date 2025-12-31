import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from django.http import HttpResponse
from django.urls import path
from django.shortcuts import render, redirect
from django.contrib import messages
from django import forms
from django.utils.translation import gettext_lazy as _
from django.utils import translation

class ImportForm(forms.Form):
    excel_file = forms.FileField(label="ملف البيانات (Excel .xlsx)")

class ExportImportMixin:
    """
    Mixin to add Professional Import/Export functionality to Django Admin models.
    Supports bilingual headers (Arabic/English) based on active language.
    """
    change_list_template = "admin/import_export_changelist.html"
    exclude_fields = ['id', 'created_at', 'updated_at'] 
    
    def get_urls(self):
        urls = super().get_urls()
        my_urls = [
            path('export-data/', self.export_data_view, name='export_data'),
            path('import-data/', self.import_data_view, name='import_data'),
            path('download-template/', self.download_template_view, name='download_template'),
        ]
        return my_urls + urls

    def get_export_fields(self):
        """Returns a list of tuples (field_name, verbose_name)"""
        fields = []
        for f in self.model._meta.fields:
            if f.name not in self.exclude_fields:
                fields.append((f.name, str(f.verbose_name)))
        return fields

    def _get_header_mapping(self):
        """Returns a dictionary mapping various header names to the technical field name."""
        mapping = {}
        for f in self.model._meta.fields:
            if f.name not in self.exclude_fields:
                # Map technical name
                mapping[f.name.lower()] = f.name
                # Map verbose name
                mapping[str(f.verbose_name).lower()] = f.name
                # Map verbose name in other languages if possible (English/Arabic)
                with translation.override('ar'):
                    mapping[str(f.verbose_name).lower()] = f.name
                with translation.override('en'):
                    mapping[str(f.verbose_name).lower()] = f.name
        return mapping

    def _style_worksheet(self, ws):
        """Apply professional styling to the worksheet."""
        header_font = Font(bold=True, color="FFFFFF", size=12)
        header_fill = PatternFill(start_color="D4AF37", end_color="D4AF37", fill_type="solid")
        center_align = Alignment(horizontal="center", vertical="center")
        thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), 
                             top=Side(style='thin'), bottom=Side(style='thin'))

        for cell in ws[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center_align
            cell.border = thin_border

        for col in ws.columns:
            max_length = 0
            column_letter = get_column_letter(col[0].column)
            for cell in col:
                cell.border = thin_border
                cell.alignment = center_align
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except: pass
            ws.column_dimensions[column_letter].width = (max_length + 5)

    def download_template_view(self, request):
        """Generate a professionally styled empty Excel template with translated headers."""
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = f'attachment; filename={self.model._meta.model_name}_template.xlsx'
        
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Template"
        
        # Set RTL based on language
        if translation.get_language() == 'ar':
            ws.sheet_view.rightToLeft = True
        
        fields = self.get_export_fields()
        headers = [f[1] for f in fields] # Use verbose_names
        ws.append(headers)
        
        self._style_worksheet(ws)
        wb.save(response)
        return response

    def export_data_view(self, request):
        """Export data with translated headers and professional styling."""
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = f'attachment; filename={self.model._meta.model_name}_export.xlsx'
        
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Exported Data"
        
        if translation.get_language() == 'ar':
            ws.sheet_view.rightToLeft = True
        
        export_config = self.get_export_fields()
        headers = [f[1] for f in export_config]
        field_names = [f[0] for f in export_config]
        
        ws.append(headers)
        
        for obj in self.get_queryset(request):
            row = []
            for field in field_names:
                val = getattr(obj, field)
                if val is None: val = ""
                row.append(str(val))
            ws.append(row)
            
        self._style_worksheet(ws)
        wb.save(response)
        return response

    def import_data_view(self, request):
        if request.method == "POST":
            form = ImportForm(request.POST, request.FILES)
            if form.is_valid():
                excel_file = request.FILES["excel_file"]
                try:
                    wb = openpyxl.load_workbook(excel_file)
                    ws = wb.active
                    
                    raw_headers = [str(cell.value).strip() if cell.value else "" for cell in ws[1]]
                    header_map = self._get_header_mapping()
                    
                    # Convert file headers to model field names
                    actual_fields = []
                    for h in raw_headers:
                        if h.lower() in header_map:
                            actual_fields.append(header_map[h.lower()])
                        else:
                            actual_fields.append(None) # Unknown column
                    
                    count = 0
                    errors = []
                    
                    for row in ws.iter_rows(min_row=2, values_only=True):
                        if not any(row): continue
                        try:
                            model_data = {}
                            for i, value in enumerate(row):
                                field_name = actual_fields[i]
                                if field_name and value is not None and str(value).strip() != "":
                                    model_data[field_name] = value
                            
                            if model_data:
                                self.model.objects.create(**model_data)
                                count += 1
                        except Exception as e:
                            errors.append(f"Row Error: {e}")
                    
                    if errors:
                        messages.warning(request, f"Imported {count} items. Errors encountered: {len(errors)}")
                    else:
                        messages.success(request, f"Successfully imported {count} items! 🚀")
                        
                    return redirect("..")
                except Exception as e:
                    messages.error(request, f"File Processing Error: {e}")
        else:
            form = ImportForm()

        context = {
            "form": form,
            "title": f"استيراد بيانات {self.model._meta.verbose_name}",
            "opts": self.model._meta,
        }
        return render(request, "admin/csv_import.html", context)
